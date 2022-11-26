from asyncore import read										# Importing libraries
from datetime import datetime
from tkinter import SOLID
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from itertools import repeat
from openpyxl.styles.borders import Border, Side

from pandas import read_excel
from pyparsing import col
start_time = datetime.now()
import os


thin_border = Border(left=Side(style='thin'),                   # Defining border
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))



def find_octant(a,b,c):                                         # Function to find the octant 
	if(a>0 and b>0 and c>0):    
		return 1
	elif(a>0 and b>0 and c<0):
		return -1
	elif(a<0 and b>0 and c>0):
		return 2
	elif(a<0 and b>0 and c<0):
		return -2
	elif(a<0 and b<0 and c>0):
		return 3
	elif(a<0 and b<0 and c<0):
		return -3
	elif(a>0 and b<0 and c>0):
		return 4
	elif(a>0 and b<0 and c<0):
		return -4


def octant_analysis(mod=5000):

	def process_file(s,mod):										# Function to process individual file
		in_file='input/'+s											# Path to file
		df=read_excel(in_file)										# Reading file in dataframe
		wb=Workbook()												# Starting a new workbook
		myworkbook=wb.active													
		lst=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']
		for i in range(11):											# Writing header of file 
			myworkbook.cell(row=2,column=i+1).value=lst[i]	
		octant=[]
		u_avg=df['U'].mean()                                        # Finding average of u,v and w                 
		v_avg=df['V'].mean()
		w_avg=df['W'].mean()
		
		myworkbook.cell(row=1,column=14).value='Overall Octant Count'
		myworkbook.cell(row=1,column=45).value='Longest Subsequence Length'
		myworkbook.cell(row=1,column=49).value='Longest Subsquence Length with Range'

		myworkbook.cell(row=3,column=5).value=round(u_avg,3)
		myworkbook.cell(row=3,column=6).value=round(v_avg,3)
		myworkbook.cell(row=3,column=7).value=round(w_avg,3)

		for i in df.index:
			myworkbook.cell(row=i+3,column=1).value=df['T'][i]
			myworkbook.cell(row=i+3,column=2).value=df['U'][i]
			myworkbook.cell(row=i+3,column=3).value=df['V'][i]
			myworkbook.cell(row=i+3,column=4).value=df['W'][i]
			myworkbook.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
			myworkbook.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
			myworkbook.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
			myworkbook.cell(row=i+3,column=11).value=find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
			octant.append(find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
		
		def octant_range_names(mod=5000):
			octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
			dic={}                                                          # creating dictionary for mapping 
			my_dic={}                                                      # Creating dictionary with opposite key value pair than 'dic'
			
			for i in range(0,4):                                            # dic[1]=0,dic[-1]=-1,...
				dic[i+1]=2*i+1-1                                            # my_dic[0]=1,my_dic[1]=-1,...
				dic[-(i+1)]=2*(i+1)-1
				my_dic[2*i+1-1]=i+1
				my_dic[2*(i+1)-1]=-(i+1)

			def find_rank_of_list(lst):                                     # Function to find the rank list from count values of all octants
				temp_lst=lst.copy()
				temp_lst.sort(reverse=True)
				res=[]

				for i in lst:
					for j in range(0,8):
						if(i==temp_lst[j]):
							res.append(j+1)
							break
				return res                                                  # Returning the ranked list
			
			def find_1st_rank(lst):                                         # Finding the octant which has rank 1 in the given rank list
				for i in range(8):
					if(lst[i]==1):
						return my_dic[i]

			def count_rank1(lst,x):                                         # Finding the count of rank 1 in the rank 1 mod values of octant x
				sum=0
				for i in lst:
					if(x==i):
						sum+=1
				return sum                                                  # Return the count
			
			my_matr=[]                                                  # Matrix to store rank list for different mod values
			rank1_list=[]                                                   # List to store the octants which have rank 1 in different mod ranges and overall
			myworkbook=wb.active
			myworkbook['M4']='Mod '+str(mod)                                           # Putting the string 'User Input' at its specified place

			matrix=[]                                                       # 2-d matrix for storing octants within ranges
			count=[0]*9                                                     # Creating a list for storing elements of 9 columns

			count[0]='Octant ID'                                            # Storing header list in 'count' list

			for i in range(0,4):
				count[2*i+1]=(i+1)
				count[2*(i+1)]=-(i+1)
			matrix.append(count)                                            # Appending header list in matrix
			for i in range(13,22):                                          # Writing header list in worksheet
				myworkbook.cell(row=3,column=i+1).value=count[i-13]
				myworkbook.cell(row=3,column=i+1).border=thin_border
				if(i>13):
					myworkbook.cell(row=3,column=i+9).value='Rank Octant '+str(count[i-13])
					myworkbook.cell(row=3,column=i+9).border=thin_border
			myworkbook.cell(row=3,column=31).value='Rank1 Octant ID'
			myworkbook.cell(row=3,column=32).value='Rank1 Octant Name'
			myworkbook.cell(row=3,column=31).border=thin_border
			myworkbook.cell(row=3,column=32).border=thin_border
			count=[0]*9                                                     # Resetting values in list 'count'
		
			for i in octant:                                                # Finding total count of values in different octants
				if(i==1):
					count[1]=count[1]+1
				elif(i==-1):
					count[2]=count[2]+1
				elif(i==2):
					count[3]=count[3]+1
				elif(i==-2):
					count[4]=count[4]+1
				elif(i==3):
					count[5]=count[5]+1
				elif(i==-3):
					count[6]=count[6]+1
				elif(i==4):
					count[7]=count[7]+1
				elif(i==-4):
					count[8]=count[8]+1
			yellow = "00FFFF00"
			count[0]='Overall Count'                                        # Creating overall count row
			matrix.append(count)                                           
			for i in range(13,22):                                          # Writing overall count in worksheet
				myworkbook.cell(row=4,column=i+1).value=count[i-13]
				myworkbook.cell(row=4,column=i+1).border=thin_border
			count.pop(0)                                                    # Removing the header from list
			rank=find_rank_of_list(count)                                   # Find the rank list 
			rank1_list.append(find_1st_rank(rank))                          # Finding the rank 1 octant and appending in rank1_list
			my_matr.append(rank)                                        # Appending rank list in the matrix
			for i in range(8):                                              # Writing overall count in worksheet
				myworkbook.cell(row=4,column=23+i).value=my_matr[0][i]
				myworkbook.cell(row=4,column=23+i).border=thin_border
				if(my_matr[0][i]==1):
					myworkbook.cell(row=4,column=23+i).fill=PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
			myworkbook.cell(row=4,column=31).value=rank1_list[0]
			myworkbook.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
			myworkbook.cell(row=4,column=31).border=thin_border
			myworkbook.cell(row=4,column=32).border=thin_border
									
			
			
			n=len(octant)                                                   # Finding the number of points given in the input
			count=[0]*9                                                     # Resetting the values in the list 'count'
			k=0                                                             # Variable to keep track of the index of data we are on
			j=4                                                             # Variable to keep track of row in worksheet
			for i in octant:                                                # Counting number of values in different octants in mod range
				if(i==1):
					count[1]=count[1]+1
				elif(i==-1):
					count[2]=count[2]+1
				elif(i==2):
					count[3]=count[3]+1
				elif(i==-2):
					count[4]=count[4]+1
				elif(i==3):
					count[5]=count[5]+1
				elif(i==-3):
					count[6]=count[6]+1
				elif(i==4):
					count[7]=count[7]+1
				elif(i==-4):
					count[8]=count[8]+1
				k=k+1                                                       # Incrementing the index tracking variable
				if(k%mod==1):                                               # Processing the mod values in the range and storing them in the list 'count'
					count[0]=str(k-1)+'-'                       
				elif(k%mod==0 or k==n):
					count[0]=count[0]+str(k-1)                              # Here count[0]-> represents the range and further elements of count represents the count in different octants
					for i in range(13,22):                                  # Writing the mod count of octant in worksheet
						myworkbook.cell(row=j+1,column=i+1).value=count[i-13]
						myworkbook.cell(row=j+1,column=i+1).border=thin_border
					count.pop(0)                                            # Removing the header from list
					rank=find_rank_of_list(count)                           # Find the rank list 
					rank1_list.append(find_1st_rank(rank))                  # Finding the rank 1 octant and appending in rank1_list
					my_matr.append(rank)                                # Appending rank list in the matrix
					
					for i in range(8):                                                  # Writing the columns of rank, rank1 and octant_name in the worksheet
						myworkbook.cell(row=j+1,column=23+i).value=my_matr[j-3][i]
						myworkbook.cell(row=j+1,column=23+i).border=thin_border
						if(my_matr[j-3][i]==1):
							myworkbook.cell(row=j+1,column=23+i).fill=PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
					
					myworkbook.cell(row=j+1,column=31).value=rank1_list[j-3]
					myworkbook.cell(row=j+1,column=32).value=octant_name_id_mapping[str(rank1_list[j-3])]
					myworkbook.cell(row=j+1,column=31).border=thin_border
					myworkbook.cell(row=j+1,column=31).border=thin_border
					
					j=j+1                                                   # Incrementing row
					matrix.append(count)
					count=[0]*9                                             # Resetting count of values in different octants    
													
			rank1_list.pop(0)                                               # Removing the overall rank1 octant
			
			myworkbook.cell(row=12,column=29).value='Octant ID'                     # Writing the header of table of count of rank1 mod values
			myworkbook.cell(row=12,column=30).value='Octant Name'
			myworkbook.cell(row=12,column=31).value='Count of Rank 1 Mod Values'
			myworkbook.cell(row=12,column=29).border=thin_border                 
			myworkbook.cell(row=12,column=30).border=thin_border
			myworkbook.cell(row=12,column=31).border=thin_border
			
			for i in range(8):                                              # Writing the table of count of rank1 mod values
				myworkbook.cell(row=13+i,column=29).value=my_dic[i]
				myworkbook.cell(row=13+i,column=30).value=octant_name_id_mapping[str(my_dic[i])]
				myworkbook.cell(row=13+i,column=31).value=count_rank1(rank1_list,my_dic[i])
				myworkbook.cell(row=13+i,column=29).border=thin_border
				myworkbook.cell(row=13+i,column=30).border=thin_border
				myworkbook.cell(row=13+i,column=31).border=thin_border
		
		def octant_longest_subsequence_count_with_range():

			r=['Count','Longest Subsequence Length','Count']                # Header list
			for i in range(3):                                              # Writing header of table to worksheet
				myworkbook.cell(row=3,column=45+i).value=r[i] 
				myworkbook.cell(row=3,column=45+i).border=thin_border    


			octants=[]
			for i in range(2,10,2):                                         # Writing octants on leftmost column of the table
				myworkbook.cell(row=i+2,column=45).value=i//2
				octants.append(i//2)
				myworkbook.cell(row=i+3,column=45).value=-(i//2) 
				octants.append(-i//2)
				myworkbook.cell(row=i+2,column=45).border=thin_border
				myworkbook.cell(row=i+3,column=45).border=thin_border                                
			
			dic={}                                                          # creating dictionary for mapping 
			for i in range(0,4):                                            
				dic[i+1]=2*i+1-1
				dic[-(i+1)]=2*(i+1)-1
					
			count=[0]*8                                                     # List for storing number of longest subsequence
			longest_length=[0]*8                                            # List for storing length of longest subsequence
			prev=octant[0]
			l=1                                                             # Length of current octant
			n=len(octant)
			temp=[0]                                                        # Temporary variable to store range
			ranges= [[] for x in repeat(None, 8)]                           # Empty list of list to store ranges for different octants

			for i in range(1,n+1):                                          # Loop for finding number and length of longest subsequence
				if(i==n):                                                   # IF last is reached process the whole
					if(longest_length[dic[prev]]<l):                        
						longest_length[dic[prev]]=l
						count[dic[prev]]=1
						temp.append(df['T'][i-1])                        # Writing ending range in temp
						ranges[dic[prev]].clear()                           # Clearing range for that octant because current longest length is small
						ranges[dic[prev]].append(temp)                      # Writing longest range for specific octant
					elif(longest_length[dic[prev]]==l):
						count[dic[prev]]+=1
						temp.append(df['T'][i-1])
						ranges[dic[prev]].append(temp)                      # Appending more ranges to the octant
				elif(prev==octant[i]):                                      # If prev and current values are same, increase current length by 1
					l+=1
				else:                                                       # Else process the previous octant values and start with new octant
					if(longest_length[dic[prev]]<l):
						longest_length[dic[prev]]=l
						count[dic[prev]]=1
						ranges[dic[prev]].clear()                           # Clearing range for that octant because current longest length is small
						temp.append(df['T'][i-1])                        # Writing ending range in temp
						ranges[dic[prev]].append(temp)                      # Writing longest range for specific octant
					elif(longest_length[dic[prev]]==l):
						count[dic[prev]]+=1
						temp.append(df['T'][i-1])
						ranges[dic[prev]].append(temp)                      # Appending more ranges to the octant
					temp=[df['T'][i]]                                    # Writing starting of range in temp variable
					l=1
					prev=octant[i]                                          # Updating previous octant for next octant

			
			for i in range(2,10):                                           # Writing the number and length of longest subsequence in table
				myworkbook.cell(row=i+2,column=46).value=longest_length[i-2]
				myworkbook.cell(row=i+2,column=47).value=count[i-2]
				myworkbook.cell(row=i+2,column=46).border=thin_border
				myworkbook.cell(row=i+2,column=47).border=thin_border
			k=2                                                             # Variable to keep track of row in worksheet
			myworkbook.cell(row=k+1,column=49).value='Octant ###'                          # Writing heading of table
			myworkbook.cell(row=k+1,column=50).value='Longest Subsequence Length'
			myworkbook.cell(row=k+1,column=51).value='Count'
			myworkbook.cell(row=k+1,column=49).border=thin_border
			myworkbook.cell(row=k+1,column=50).border=thin_border
			myworkbook.cell(row=k+1,column=51).border=thin_border
			
			k+=2
			for i in range(8):
				myworkbook.cell(row=k,column=49).value=octants[i]                   # Writing contents of table-1 for each octant
				myworkbook.cell(row=k,column=50).value=longest_length[i]
				myworkbook.cell(row=k,column=51).value=count[i]
				myworkbook.cell(row=k+1,column=49).value='Time'                     # Writing header of ranges in worksheet
				myworkbook.cell(row=k+1,column=50).value='From'
				myworkbook.cell(row=k+1,column=51).value='To'
				myworkbook.cell(row=k,column=49).border=thin_border                 # Adding border to cells
				myworkbook.cell(row=k,column=50).border=thin_border
				myworkbook.cell(row=k,column=51).border=thin_border
				myworkbook.cell(row=k+1,column=49).border=thin_border
				myworkbook.cell(row=k+1,column=50).border=thin_border
				myworkbook.cell(row=k+1,column=51).border=thin_border
				x=ranges[i]
				k+=2
				for j in x:
					myworkbook.cell(row=k,column=50).value=j[0]                     # Writing ranges in worksheet
					myworkbook.cell(row=k,column=51).value=j[1]
					myworkbook.cell(row=k,column=49).border=thin_border             # Adding border to cells
					myworkbook.cell(row=k,column=50).border=thin_border
					myworkbook.cell(row=k,column=51).border=thin_border
					k+=1
			
		def octant_transition_count(mod=5000):
			j=1
			n=len(octant)
			myworkbook.cell(row=j,column=35).value='Overall Transition Count'       # Writing overall transition count in worksheet
			myworkbook.cell(row=j+3,column=34).value='From'
			myworkbook.cell(row=j+1,column=36).value='To'
			j+=2
			
			matrix = [ [0]*9 for i in range(9)]                             # Creating 9*9 matrix for storing transition count values
			
			for i in range(0,4):                                            # Storing header row and header column in the matrix
				matrix[0][2*i+1]=(i+1)
				matrix[0][2*(i+1)]=-(i+1)
			for i in range(0,9):
				matrix[i][0]=matrix[0][i]
			matrix[0][0]='Octant #'

			dic={}                                                          # creating dictionary for mapping 
			for i in range(0,4):
				dic[i+1]=2*i+1
				dic[-(i+1)]=2*(i+1)

			def find_row_col(x,y):                                          # Finding row and column of matrix from transition values
				lst=[dic[x],dic[y]]
				return lst
			
			def find_max_ele(lst):
				temp=lst.copy()
				temp.pop(0)
				large=0
				for i in temp:
					if(large<i):
						large=i
				return large

			prev=octant[0]                                              
			for i in range(1,n):                                            # Filling overall transition matrix
				lst=find_row_col(prev,octant[i])                            # lst[0]-> row and lst[1]->column of overall transition matrix
				matrix[lst[0]][lst[1]]+=1
				prev=octant[i]
			yellow = "00FFFF00"
			for i in range(0,9):                                            # Writing the overall transition matrix in worksheet
				temp_lst=matrix[i]
				large=find_max_ele(temp_lst)
				for k in range(13,22):
					myworkbook.cell(row=j+i,column=k+22).value=matrix[i][k-13]
					myworkbook.cell(row=j+i,column=k+22).border=thin_border
					if(i>0 and matrix[i][k-13]==large):
						myworkbook.cell(row=j+i,column=k+22).fill=PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
					if(i!=0 and k!=13):
						matrix[i][k-13]=0
				
			temp=n//mod+1                                                   # temp-> No. of mod transition tables
			j+=1
			for t in range(0,temp):                                         # One iteration for each mod transition table
				j+=11
				name=''
				myworkbook.cell(row=j,column=35).value='Mod Transition Count'       # Writing Table name in worksheet
				myworkbook.cell(row=j+3,column=34).value='From'
				myworkbook.cell(row=j+1,column=36).value='To'
				name=str(t*mod)+'-'
				if((t+1)*mod-1>n-1):
					name+=str(n-1)
				else:
					name+=str((t+1)*mod-1)   
				myworkbook.cell(row=j+1,column=35).value=name
				j+=2

				for i in range(t*mod,min(n-1,(t+1)*mod)):                   # Incrementing matrix cell corresponding to transition values
					lst=find_row_col(octant[i],octant[i+1])
					matrix[lst[0]][lst[1]]+=1

				for i in range(0,9):                                        # Writing the transition mod matrix in worksheet
					temp_lst=matrix[i]
					if(i>0):
						large=find_max_ele(temp_lst)
					for k in range(13,22):
						myworkbook.cell(row=j+i,column=k+22).value=matrix[i][k-13]
						myworkbook.cell(row=j+i,column=k+22).border=thin_border
						if(i>0 and matrix[i][k-13]==large):
							myworkbook.cell(row=j+i,column=k+22).fill=PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
						if(i!=0 and k!=13):
							matrix[i][k-13]=0                               # Resetting matrix for next mod iteration

		
		octant_transition_count(mod=5000)
		octant_range_names(5000)
		octant_longest_subsequence_count_with_range()
		s=s[:-5]
		file_name='output/'+s+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		wb.save(file_name)
	
	os.mkdir('output')
	input_files=os.listdir('input')
	for i in range(len(input_files)):
		process_file(input_files[i],mod)
		


##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)